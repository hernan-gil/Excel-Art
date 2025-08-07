import os
import sys
from PIL import Image
import openpyxl
from openpyxl.styles import PatternFill
import argparse

def rgb_to_hex(rgb):
    """Convierte valores RGB a formato hexadecimal para Excel"""
    return f"{rgb[0]:02X}{rgb[1]:02X}{rgb[2]:02X}"

def load_image(image_path):
    """Carga una imagen y la convierte a RGB si es necesario"""
    try:
        image = Image.open(image_path)
        # Convertir a RGB si la imagen tiene canal alpha o está en otro formato
        if image.mode != 'RGB':
            image = image.convert('RGB')
        return image
    except Exception as e:
        print(f"Error al cargar la imagen: {e}")
        return None

def get_pixel_colors(image):
    """Extrae los colores de todos los pixels de la imagen"""
    width, height = image.size
    pixel_colors = []
    
    print(f"Procesando imagen de {width}x{height} pixels...")
    
    for y in range(height):
        row_colors = []
        for x in range(width):
            # Obtener el color del pixel en la posición (x, y)
            pixel_color = image.getpixel((x, y))
            row_colors.append(pixel_color)
        pixel_colors.append(row_colors)
    
    return pixel_colors, width, height

def create_excel_with_colors(pixel_colors, width, height, output_path):
    """Crea un archivo Excel donde cada celda tiene el color del pixel correspondiente"""
    print("Creando archivo Excel...")
    
    # Crear un nuevo workbook
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "Imagen_Pixeles"
    
    # OPTIMIZACIÓN 1: Cache de estilos para evitar crear fills duplicados
    fill_cache = {}
    
    # OPTIMIZACIÓN 2: Ajustar tamaño de celdas de forma más eficiente
    # Hacer las celdas más pequeñas y cuadradas
    for col in range(1, width + 1):
        worksheet.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 1.5
    
    for row in range(1, height + 1):
        worksheet.row_dimensions[row].height = 12
    
    # OPTIMIZACIÓN 3: Procesar por lotes y reusar fills
    total_cells = width * height
    processed = 0
    
    for y in range(height):
        for x in range(width):
            rgb_color = pixel_colors[y][x]
            hex_color = rgb_to_hex(rgb_color)
            
            # Reusar fill si ya existe para este color
            if hex_color not in fill_cache:
                fill_cache[hex_color] = PatternFill(
                    start_color=hex_color, 
                    end_color=hex_color, 
                    fill_type="solid"
                )
            
            # Aplicar el color a la celda (Excel usa indexación 1-based)
            cell = worksheet.cell(row=y + 1, column=x + 1)
            cell.fill = fill_cache[hex_color]
            
            processed += 1
            
        # Mostrar progreso cada 10 filas
        if (y + 1) % 10 == 0:
            progress = (processed / total_cells) * 100
            print(f"Procesadas {y + 1}/{height} filas ({progress:.1f}% - {len(fill_cache)} colores únicos)")
    
    print(f"Total de colores únicos encontrados: {len(fill_cache)}")
    
    # OPTIMIZACIÓN 4: Guardar con configuraciones de rendimiento
    try:
        # Deshabilitar cálculos automáticos para mejorar rendimiento
        workbook.calculation.calcMode = 'manual'
        workbook.save(output_path)
        print(f"Archivo Excel guardado como: {output_path}")
        print(f"NOTA: El archivo puede tardar en abrir debido a {total_cells:,} celdas coloreadas")
        return True
    except Exception as e:
        print(f"Error al guardar el archivo Excel: {e}")
        return False

def reduce_colors(image, max_colors=256):
    """Reduce el número de colores en la imagen para mejor rendimiento"""
    print(f"Reduciendo colores a máximo {max_colors} para mejor rendimiento...")
    # Cuantizar la imagen para reducir colores
    if image.mode != 'P':
        image = image.quantize(colors=max_colors)
        image = image.convert('RGB')
    return image

def image_to_excel(image_path):
    """Función principal que convierte una imagen a Excel"""
    # Verificar que el archivo existe
    if not os.path.exists(image_path):
        print(f"Error: El archivo {image_path} no existe.")
        return False
    
    # Verificar que es una imagen válida
    valid_extensions = ['.png', '.jpg', '.jpeg', '.PNG', '.JPG', '.JPEG']
    if not any(image_path.lower().endswith(ext) for ext in valid_extensions):
        print("Error: El archivo debe ser PNG o JPG/JPEG.")
        return False
    
    print(f"Cargando imagen: {image_path}")
    
    # 1. Cargar la imagen
    image = load_image(image_path)
    if image is None:
        return False
    
    # Advertencia para imágenes muy grandes
    total_pixels = image.width * image.height
    if total_pixels > 10000:  # Más de 100x100
        print(f"La imagen tiene {total_pixels:,} pixels.")
        print("Para mejor rendimiento, se recomienda:")
        print("1. Reducir colores (más rápido)")
        print("2. Continuar sin cambios (más lento)")
        print("3. Cancelar")
        
        choice = input("Elige una opción (1/2/3): ").strip()
        if choice == '1':
            image = reduce_colors(image)
        elif choice == '3':
            print("Operación cancelada.")
            return False
        # Si elige 2, continúa sin cambios
    
    # 2. Obtener información de color de cada pixel
    pixel_colors, width, height = get_pixel_colors(image)
    
    # 3. Crear nombre del archivo de salida
    base_name = os.path.splitext(image_path)[0]
    output_path = f"{base_name}.xlsx"
    
    # 4. Crear el archivo Excel con los colores
    success = create_excel_with_colors(pixel_colors, width, height, output_path)
    
    if success:
        print(f"\n¡Conversión completada!")
        print(f"Imagen procesada: {width}x{height} pixels")
        print(f"Archivo Excel creado: {output_path}")
        print(f"\nTIPS para abrir el archivo más rápido:")
        print("- Usa una computadora con buena RAM")
        print("- Cierra otros programas antes de abrir Excel")
        print("- Ten paciencia, puede tardar 1-2 minutos en cargar")
        return True
    else:
        return False

def main():
    """Función principal con interfaz de línea de comandos"""
    parser = argparse.ArgumentParser(description='Convierte una imagen PNG/JPG a Excel con colores de pixels')
    parser.add_argument('image_path', help='Ruta a la imagen PNG o JPG')
    
    # Si no se proporcionan argumentos, pedir la ruta interactivamente
    if len(sys.argv) == 1:
        print("=== Conversor de Imagen a Excel ===")
        print("Este programa convierte una imagen PNG/JPG a un archivo Excel")
        print("donde cada celda representa un pixel con su color correspondiente.\n")
        
        image_path = input("Ingresa la ruta de la imagen: ").strip()
        if not image_path:
            print("No se proporcionó ninguna ruta.")
            return
    else:
        args = parser.parse_args()
        image_path = args.image_path
    
    # Procesar la imagen
    image_to_excel(image_path)

if __name__ == "__main__":
    main()